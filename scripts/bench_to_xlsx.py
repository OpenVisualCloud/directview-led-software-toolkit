#!/usr/bin/env python3
"""Render dvledtx benchmark results into the uncompressed-frames XLSX report.

Usage: bench_to_xlsx.py <results.csv> <output.xlsx>

Only the Intel E610-XT4 columns are populated from the measured CSV; the
Intel I225V columns carry the previously recorded reference figures.
"""
import csv
import sys
from collections import OrderedDict

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

FORMATS = OrderedDict([
    ("yuv420", "YUV420"),
    ("yuv422p10le", "YUV422P10LE"),
    ("yuv444p10le", "YUV444P10LE"),
    ("gbrp10le", "GBRP10LE"),
    ("yuv422p12le", "YUV422P12LE"),
    ("yuv444p12le", "YUV444P12LE"),
    ("gbrp12le", "GBRP12LE"),
])

GEOMETRIES = [
    (1920, 1080, 30),
    (1920, 1080, 60),
    (2560, 1440, 30),
    (2560, 1440, 60),
    (3840, 2160, 30),
    (3840, 2160, 60),
]

# Previously recorded Intel I225V figures: (w, h, fps, fmt) -> (fps, Mbps)
I225V = {
    (1920, 1080, 30, "yuv420"): ("29.9-30", 784.56),
    (1920, 1080, 30, "yuv422p10le"): ("29.9-30", 1306.87),
    (1920, 1080, 30, "yuv444p10le"): ("29.9-30", 1959.41),
    (1920, 1080, 30, "gbrp10le"): ("29.9-30", 1959.43),
    (1920, 1080, 30, "yuv422p12le"): ("29.9-30", 1567),
    (1920, 1080, 30, "yuv444p12le"): ("29.9-30", 2351),
    (1920, 1080, 30, "gbrp12le"): ("29.9-30", 2351),
    (1920, 1080, 60, "yuv420"): ("59.9-60", 1560.13),
    (1920, 1080, 60, "yuv422p10le"): ("~56", 2449.91),
    (1920, 1080, 60, "yuv444p10le"): ("~37.5", 2449.91),
    (1920, 1080, 60, "gbrp10le"): ("~37.5", 2449.91),
    (1920, 1080, 60, "yuv422p12le"): ("46.8", 2449.9),
    (1920, 1080, 60, "yuv444p12le"): ("31.2", 2449.9),
    (1920, 1080, 60, "gbrp12le"): ("31.2", 2449.9),
}

THIN = Side(style="thin", color="B0B0B0")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
HDR_FILL = PatternFill("solid", fgColor="1F4E79")
SUB_FILL = PatternFill("solid", fgColor="DDEBF7")
GEO_FILL = PatternFill("solid", fgColor="F2F2F2")
OK_FILL = PatternFill("solid", fgColor="E2EFDA")
BAD_FILL = PatternFill("solid", fgColor="FCE4D6")
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)


def load(path):
    rows = {}
    with open(path, newline="") as fh:
        for r in csv.DictReader(fh):
            try:
                key = (int(r["width"]), int(r["height"]),
                       int(r["fps_target"]), r["fmt"])
            except (KeyError, ValueError):
                continue
            if int(r.get("samples") or 0) == 0:
                continue
            rows[key] = r  # later runs win
    return rows


def main():
    if len(sys.argv) != 3:
        print(__doc__)
        return 2
    csv_path, xlsx_path = sys.argv[1], sys.argv[2]
    data = load(csv_path)

    wb = Workbook()
    ws = wb.active
    ws.title = "Uncompressed frames"

    ws["A1"] = "Uncompressed frames"
    ws["A1"].font = Font(bold=True, size=13)

    # Header block (rows 3-4)
    ws.merge_cells("A3:A4")
    ws["A3"] = "Resolution"
    ws.merge_cells("B3:B4")
    ws["B3"] = "Format"
    ws.merge_cells("C3:D3")
    ws["C3"] = "Intel I225V"
    ws.merge_cells("E3:G3")
    ws["E3"] = "Intel E610-XT4"
    ws["C4"] = "Achieved FPS"
    ws["D4"] = "Bandwidth (Mbps)"
    ws["E4"] = "Achieved FPS"
    ws["F4"] = "Bandwidth (Mbps)"
    ws["G4"] = "Peak Bandwidth (Mbps)"

    for row in (3, 4):
        for col in range(1, 8):
            c = ws.cell(row=row, column=col)
            c.font = Font(bold=True, color="FFFFFF" if row == 3 else "1F4E79")
            c.fill = HDR_FILL if row == 3 else SUB_FILL
            c.alignment = CENTER
            c.border = BORDER

    r = 5
    for w, h, fps in GEOMETRIES:
        first = r
        for fmt_key, fmt_label in FORMATS.items():
            ws.cell(row=r, column=2, value=fmt_label)
            ref = I225V.get((w, h, fps, fmt_key))
            if ref:
                ws.cell(row=r, column=3, value=ref[0])
                ws.cell(row=r, column=4, value=ref[1])

            rec = data.get((w, h, fps, fmt_key))
            if rec:
                f_avg = float(rec["fps_avg"])
                f_min = float(rec["fps_min"])
                f_max = float(rec["fps_max"])
                mbps = float(rec["mbps_avg"])
                label = (f"{f_avg:.1f}" if (f_max - f_min) < 0.15
                         else f"{f_min:.1f}-{f_max:.1f}")
                fc = ws.cell(row=r, column=5, value=label)
                bc = ws.cell(row=r, column=6, value=round(mbps, 2))
                bc.number_format = "0.00"
                peak = rec.get("mbps_max")
                if peak not in (None, ""):
                    pc = ws.cell(row=r, column=7, value=round(float(peak), 2))
                    pc.number_format = "0.00"
                fc.fill = OK_FILL if f_avg >= fps * 0.98 else BAD_FILL
            for col in range(1, 8):
                cc = ws.cell(row=r, column=col)
                cc.alignment = CENTER
                cc.border = BORDER
            r += 1

        ws.merge_cells(start_row=first, start_column=1, end_row=r - 1, end_column=1)
        gc = ws.cell(row=first, column=1, value=f"{w} x {h}\n{fps} FPS")
        gc.alignment = CENTER
        gc.font = Font(bold=True)
        gc.fill = GEO_FILL

    for col, width in zip("ABCDEFG", (16, 18, 15, 18, 15, 18, 20)):
        ws.column_dimensions[col].width = width
    ws.freeze_panes = "C5"

    # Raw measurement sheet
    ws2 = wb.create_sheet("Raw")
    with open(csv_path, newline="") as fh:
        for i, row in enumerate(csv.reader(fh), start=1):
            for j, val in enumerate(row, start=1):
                try:
                    val = float(val) if "." in val else int(val)
                except (ValueError, TypeError):
                    pass
                cell = ws2.cell(row=i, column=j, value=val)
                if i == 1:
                    cell.font = Font(bold=True)
    for j in range(1, 15):
        ws2.column_dimensions[get_column_letter(j)].width = 15
    ws2.freeze_panes = "A2"

    wb.save(xlsx_path)
    print(f"wrote {xlsx_path} ({len(data)} measured cases)")
    return 0


if __name__ == "__main__":
    sys.exit(main())
